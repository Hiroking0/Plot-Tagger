# -*- coding: utf-8 -*-
# SPDX-License-Identifier: Apache-2.0
# Copyright © 2025 Hiroki Fujisato

"""
Created on Mon Jan 15 12:20:24 2025

@author: Hiroki Fujisato
"""

import pandas as pd
from tabulate import tabulate
from hdf5_handle import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

current_path = os.path.dirname(os.path.abspath(__file__))
# Define the HDF5 file path
hdf5_file_path = os.path.join(current_path,'tag_data.h5')

with h5py.File(hdf5_file_path, 'r+') as f:
   df  = loadH5Table(f,'dataset').applymap(lambda x: x.decode('utf-8') if isinstance(x, bytes) else x)

# Save the DataFrame to an Excel file
# Get the directory where the script is running
current_script_directory = os.path.dirname(os.path.abspath(__file__))

# Define the Excel file path (save it in the same directory as the script)
excel_file_path = os.path.join(current_script_directory, 'tag_id.xlsx')

# Save the DataFrame to the Excel file
df.to_excel(excel_file_path, index=False)

# Open the Excel file to apply the style
wb = load_workbook(excel_file_path)
ws = wb.active

# Define the fill color for "New" status
grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
# Iterate over the rows and highlight rows where 'Status' is 'New'
# Iterate over the rows and highlight rows based on the 'Status' column
for row in range(2, len(df) + 2):  # Start from 2 to skip the header row
    status_cell = ws.cell(row=row, column=df.columns.get_loc('Status') + 1)
    
    if status_cell.value == "New":
        # Apply green color to the entire row if the status is "New"
        for col in range(1, len(df.columns) + 1):
            ws.cell(row=row, column=col).fill = green_fill
    elif status_cell.value == "Replaced":
        # Apply grey color to the entire row if the status is "Replaced"
        for col in range(1, len(df.columns) + 1):
            ws.cell(row=row, column=col).fill = grey_fill
wb.save(excel_file_path)
